import pathlib
from typing import Any, List

import openpyxl.reader.excel
import openpyxl.utils
import pandas

from processing.classes.Config import Config


class BaseConfigWriter:
    def __init__(
        self,
        filename: str,
    ):
        self.__verify_and_set_path(filename)

        self.__load()

        self.__config = Config()
        self.__columns = self.__config.get_columns()

    def __verify_and_set_path(self, filename: str):
        path = pathlib.Path(filename).absolute()

        if not path.exists():
            raise FileNotFoundError(f'{path}')

        self.__path = path

    def __load(self):
        self.__workbook = openpyxl.reader.excel.load_workbook(self.__path)
        self.__worksheet = self.__workbook["Sheet1"]

    def get_column_letter(self, column_name):
        return self.__columns[column_name]['letter']

    def save(self):
        self.__workbook.save(self.__path)

    def write_column(self, column_name: str, values: List[Any]):
        try:
            column_letter = self.get_column_letter(column_name)

            dataframe = pandas.DataFrame(
                {
                    column_letter: [column_name, None, *values]
                }
            )

            print(f"Config: Writing column {column_name} ({column_letter})")

            for i, row in dataframe.iterrows():
                cell = f'{column_letter}%d' % (i + 1)
                self.__worksheet[cell] = row[0]
        except KeyError:
            # Excel column not found
            # User does not want this indicator
            pass

    def write(self):
        """
        Implement in children.

        Call `.save()` method at the end.
        """
        pass
