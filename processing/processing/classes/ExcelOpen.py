from typing import Dict, List, Union

from openpyxl.reader.excel import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from processing.utils.add_unique_value_only_to_array import \
    add_unique_value_only_to_array


class ExcelOpen:
    """The ExcelOpen configuration file using `openpyxl`.

    Cherry-pick cells, append and write to current file.

    Attributes:
        __path: The path to configuration file.
        __workbook: The loaded workbook.
        __worksheet: The current worksheet.
        __max_row: The current sheet's max row count.

        columns: The dictionary of all Excel columns.
            Keys are column cell content.
            Values are dictionaries made of:
                - The `letter` key with its corresponding value.
                - The `number` key with its corresponding value.

        meta_properties: The list of meta properties.
        meta_values: The list of meta values.
        meta_values_uniques: The list of meta values with no duplicates.
    """
    __path: str
    __workbook: Workbook
    __worksheet: Worksheet
    __max_row: int
    columns: Dict[str, Dict[str, Union[str, int]]]
    meta_properties: List[str]
    meta_values: List[List[str]]
    meta_values_uniques: List[List[str]]

    def __init__(
        self,
        path: str,
    ) -> None:
        self.__path = path

        self.__open()
        self.__get_columns()
        self.__get_meta_properties_and_values()
        self.__close()

    def __open(self):
        self.__workbook = load_workbook(self.__path)
        self.__worksheet = self.__workbook["Sheet1"]
        self.__max_row = self.__worksheet.max_row

    def __close(self):
        self.__workbook.save(self.__path)

    def __get_columns(self):
        self.columns = {
            cell.value: {
                'letter': get_column_letter(cell.column),
                'number': cell.column - 1
            } for cell in self.__worksheet[1] if cell.value
        }

    def __get_meta_properties_and_values(self):
        self.meta_properties = []
        self.meta_values = []
        self.meta_values_uniques = []

        for column_key in self.columns.keys():
            if 'files' not in column_key:
                continue
            if 'files_start' in column_key:
                continue
            if 'files_tags' in column_key:
                continue
            if 'files_site' in column_key:
                continue
            if column_key == 'files':
                continue
            if column_key == 'files (no suffix)':
                continue

            # trim `files_`
            meta_property = "_".join(column_key.split('_')[1:])
            self.meta_properties.append(meta_property)

            meta_property_column = self.columns[column_key]['letter']

            meta_values, meta_values_uniques = self.__get_values_from_column(
                meta_property_column
            )

            self.meta_values.append(meta_values)
            self.meta_values_uniques.append(meta_values_uniques)

    def __get_values_from_column(self, column):
        results = []
        uniques = []

        for row in range(3, self.__max_row - 1):
            cell_name = "{}{}".format(column, row)
            value = self.__worksheet[cell_name].value

            if isinstance(value, int):
                value = str(value)

            results.append(value)
            add_unique_value_only_to_array(uniques, value)

        return results, uniques
