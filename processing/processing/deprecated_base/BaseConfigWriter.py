import pathlib
from pathlib import Path
from typing import Any, Dict, List, Union

import openpyxl.reader.excel
import openpyxl.utils
import pandas
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from processing.deprecated_classes.Config import Config


class BaseConfigWriter:
    """The configuration writer base to inherit from.

    Children must implement `.write()` method.

    Attributes:
        __config: The user configuration.
        __columns: The Excel columns.
        __path: The path to Excel configuration file.
        __workbook: The Workbook object loaded with `openpyxl`.
        __worksheet: The currently selected worksheet.
    """
    __config: Config
    __columns: Dict[str, Dict[str, Union[str, int]]]
    __path: Path
    __workbook: Workbook
    __worksheet: Worksheet

    def __new__(
        cls,
        filename: str,
        *args,
        **kwargs,
    ):
        if cls is BaseConfigWriter:
            raise TypeError(
                f"only children of '{cls.__name__}' may be instantiated"
            )

        # noinspection PyArgumentList
        return object.__new__(cls, *args, **kwargs)

    def __init__(
        self,
        filename: str,
    ):
        self.__verify_and_set_path(filename)

        self.__load()

        self.__config = Config()
        self.__columns = self.__config.get_columns()

    def __verify_and_set_path(self, filename: str):
        path = pathlib.Path(filename).absolute()

        if not path.exists():
            raise FileNotFoundError(f'{path}')

        self.__path = path

    def __load(self):
        self.__workbook = openpyxl.reader.excel.load_workbook(self.__path)
        self.__worksheet = self.__workbook["Sheet1"]

    def get_column_letter(self, column_name):
        return self.__columns[column_name]['letter']

    def save(self):
        self.__workbook.save(self.__path)

    def write_column(self, column_name: str, values: List[Any]):
        try:
            column_letter = self.get_column_letter(column_name)

            dataframe = pandas.DataFrame(
                {
                    column_letter: [column_name, None, *values]
                }
            )

            print(f"Config: Writing column {column_name} ({column_letter})")

            for i, row in dataframe.iterrows():
                index = int(str(i)) + 1
                cell = f'{column_letter}%d' % index
                self.__worksheet[cell] = row[0]
        except KeyError:
            # Excel column not found
            # User does not want this indicator
            pass

    def write(self):
        """Implement in children.

        Call `.save()` method at the end.
        """
        pass
